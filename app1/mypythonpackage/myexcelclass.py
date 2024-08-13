import os

import openpyxl
import pandas
from django.shortcuts import render
import datetime, time
import xlsxwriter
from openpyxl.reader.excel import load_workbook
import pandas as pd

base_path = os.path.dirname(os.path.abspath(__file__)).split("project1")[0].replace("\\","/")

year = datetime.datetime.today().year
year_list = [year - 2, year - 1, year, year + 1, year + 2]

month_dict = {
    1: 'January',
    2: 'February',
    3: 'March',
    4: 'April',
    5: 'May',
    6: 'June',
    7: 'July',
    8: 'August',
    9: 'September',
    10: 'October',
    11: 'November',
    12: 'December',
}
month_ = month_dict[datetime.datetime.today().month]
listofmonths = [month_dict[a] for a in range(1, 13)]


def get_value_from_web(commodity, filename):
    # month = commodity['commodity_1'][2]
    month = datetime.datetime.today().month

    if filename == "":
        print("coming till here file empty")
        return "File name cannot be empty"
    path = base_path + 'project1/app/Excelfiles/'
    filename = path + filename
    
    """
        # if file is not created
        commodity_name = commodity['commodity_1'][0]
        price = commodity['commodity_1'][1]
        date = commodity['commodity_1'][2]
        place = commodity['commodity_1'][3]
        category = commodity['commodity_1'][4]
        # date = str(datetime.date.today())

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.title = "Commodities_details"
        sheet['A1'] = 'Commodity Name'
        sheet['B1'] = 'Price'
        sheet['C1'] = 'Date'
        sheet['D1'] = 'Place'
        sheet['E1'] = 'Category'

        workbook.save(filename)

        workbook = openpyxl.load_workbook(filename)
        sheet = workbook['Commodities_details']
        row = (commodity_name, int(price), date, place, category)
        sheet.append(row)
        workbook.save(filename)
    """

    try:
        commodity_name = commodity['commodity_1'][0]
        price = commodity['commodity_1'][1]
        date = commodity['commodity_1'][2]
        place = commodity['commodity_1'][3]
        # date = str(datetime.date.today())
        category = commodity['commodity_1'][4]

        new_record_data = [commodity_name, price, date, place, category]
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Get all existing IDs
        existing_ids = set()
        for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
            existing_ids.add(row[5])
        # print(existing_ids)
        # Check if there are any deleted IDs
        deleted_ids = set(range(1, len(existing_ids) + 2)) - existing_ids

        # Use a deleted ID if available, otherwise generate a new ID
        if deleted_ids:
            new_uniq_id = deleted_ids.pop()
            # print(new_uniq_id)
        else:
            new_uniq_id = len(existing_ids) + 1

        # Add the new recor

        # row_number = sheet.max_row + 1
        sheet.append([*new_record_data, new_uniq_id])
        workbook.save(filename)

        return f" {new_record_data} record inserted successfully."

    except Exception as e:
        return f"An error occurred: {str(e)}"


def verify_unique_id(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Get the active worksheet
    ws = wb.active

    # Check if the ID column contains unique values
    ids = set()

    for row in ws.iter_rows(min_row=1, min_col=1, max_col=6, values_only=True):
        # print(row[0],ids)
        if row[0] in ids:
            # print("Duplicate IDs found")
            return "Unique IDs Problem (IDs are not matching)"
        ids.add(row[0])

    # print("All IDs are unique")


def crtnewbillfile(filename):
    path = base_path + 'project1/app/Excelfiles/'
    full_file_loca = path + filename + '.xlsx'
    # print(full_file_loca)
    # print(not os.path.isfile(full_file_loca), os.path.isfile(full_file_loca))

    if not os.path.isfile(full_file_loca):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Commodities_details"
        sheet['A1'] = 'Commodity Name'
        sheet['B1'] = 'Price'
        sheet['C1'] = 'Date'
        sheet['D1'] = 'Place'
        sheet['E1'] = 'Category'
        sheet['F1'] = 'Unique IDs'

        workbook.save(full_file_loca)

        workbook = openpyxl.load_workbook(full_file_loca)
        workbook.save(full_file_loca)
        return f"{filename}.xlsv file created Successfully"
    else:
        return f"{filename}.xlsx file already exist"


def getlistofexlfiles():
    folder_path = base_path + 'project1/app1/Excelfiles'
    files = os.listdir(folder_path)
    file_creation_times = [(file, os.path.getctime(os.path.join(folder_path, file))) for file in files]
    sorted_files = sorted(file_creation_times, key=lambda x: x[1])

    arr = []
    for file, creation_time in sorted_files:
        arr.append(file)
    return arr


def load_filedata_to_templeate(file):
    path = base_path + 'project1/app1/Excelfiles/' + file
    workbook = load_workbook(path)
    worksheet = workbook.active
    filedata = []
    for row in worksheet.iter_rows(values_only=True):
        if row[0] is not None:
            filedata.append(row)
    df = pd.read_excel(path)
    # Remove empty records (rows with all NaN values)
    df.dropna(how='all', inplace=True)

    return filedata


def edit_record(file, record_id, new_values):

    try :

        file_path = base_path + 'project1/app/Excelfiles/' + file
        workbook = openpyxl.load_workbook(file_path)

        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[5].value == int(record_id):
                for col_index, new_value in enumerate(new_values, start=1):
                    row[col_index-1].value = new_value
                break
        workbook.save(file_path)
        return "Record Edited and Saved successfully"

    except Exception as e:
        return "Something is Wrong with Data Formats"

def delete_record(file, record_id):
    try:
        file_path = base_path + 'project1/app/Excelfiles/' + file
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Check if the unique ID matches (assuming unique ID is in the 6th column)
                if row[5].value == int(record_id):
                    sheet.delete_rows(row[0].row)
        wb.save(file_path)

        return "Record Deleted Successfully"
    except Exception as e:
        return "Something is Wrong with Data Formats"


